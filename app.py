"""App Flask - Exógena DIAN (Postgres + Auth)."""
from __future__ import annotations

import io
import os
import shutil
import tempfile
import warnings
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from flask import (
    Flask, flash, redirect, render_template, request, send_file, session, url_for,
)
from flask_login import login_required
from openpyxl.cell import WriteOnlyCell
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from sqlalchemy.engine import make_url
from sqlalchemy.exc import ArgumentError
from werkzeug.utils import secure_filename

from auth import bp as auth_bp, login_manager
from core import directorio, helpers, parser, registry
from db import db, migrar_codigos_territoriales, usuario_actual_id
from generators import GENERADORES
from generators.base import ContextoInformante, _config, preparar_dataframe_generacion


load_dotenv()

BASE = Path(__file__).resolve().parent
UPLOADS = BASE / "uploads"
UPLOADS.mkdir(exist_ok=True)

XLSX_STREAMING_CELL_THRESHOLD = int(os.environ.get("XLSX_STREAMING_CELL_THRESHOLD", "250000"))
XLSX_SPOOL_MAX_SIZE = int(os.environ.get("XLSX_SPOOL_MAX_SIZE", str(8 * 1024 * 1024)))


def _normalizar_database_url(raw: str | None) -> str:
    db_url = (raw or "").strip().strip('"').strip("'")
    if not db_url:
        return "sqlite:///dev.db"

    # Render/Heroku usan postgres:// pero SQLAlchemy 2.0 quiere postgresql://
    if db_url.startswith("postgres://"):
        db_url = db_url.replace("postgres://", "postgresql+psycopg://", 1)
    elif db_url.startswith("postgresql://"):
        db_url = db_url.replace("postgresql://", "postgresql+psycopg://", 1)

    try:
        make_url(db_url)
    except ArgumentError as exc:
        raise RuntimeError(
            "DATABASE_URL invalida. En Coolify pon solo el valor de la URL, sin "
            "'DATABASE_URL=', sin comillas, y con caracteres especiales del password "
            "codificados para URL."
        ) from exc
    return db_url


def crear_app() -> Flask:
    app = Flask(__name__)

    db_url = _normalizar_database_url(os.environ.get("DATABASE_URL"))

    app.config["SQLALCHEMY_DATABASE_URI"] = db_url
    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {"pool_pre_ping": True, "pool_recycle": 300}
    app.secret_key = os.environ.get("SECRET_KEY", "dev-not-secret-change-me")
    app.config["MAX_CONTENT_LENGTH"] = 256 * 1024 * 1024
    app.config["PERMITIR_REGISTRO_PUBLICO"] = (
        os.environ.get("PERMITIR_REGISTRO_PUBLICO", "true").lower() != "false"
    )

    db.init_app(app)
    login_manager.init_app(app)
    app.register_blueprint(auth_bp)

    with app.app_context():
        db.create_all()
        migrar_codigos_territoriales()
        registry.sembrar_catalogo_global()

    _registrar_rutas(app)
    return app


# ---------------- helpers de sesión web ----------------

def _user_dir(sub: str) -> Path:
    uid = usuario_actual_id()
    d = UPLOADS / f"u{uid}"
    d.mkdir(exist_ok=True)
    if sub:
        d = d / sub
        d.mkdir(exist_ok=True, parents=True)
    return d


def _cargar_df():
    p = _user_dir("") / "data.parquet"
    if not p.exists():
        return None
    return pd.read_parquet(p)


def _guardar_df(df: pd.DataFrame):
    p = _user_dir("") / "data.parquet"
    df.to_parquet(p, index=False)


def _informante_actual() -> dict[str, str] | None:
    informantes = session.get("informantes", [])
    if not informantes:
        return None
    primero = informantes[0] or {}
    nit = str(primero.get("nit", "")).strip()
    nombre = str(primero.get("nombre", "")).strip()
    if not nit:
        return None
    return {"nit": nit, "nombre": nombre, "tdoc": str(helpers.inferir_tipo_documento(nit, nombre))}


def _entero_form(valor, default: int) -> int:
    try:
        return int(valor)
    except (TypeError, ValueError):
        return default


# ---------------- rutas ----------------

def _registrar_rutas(app: Flask) -> None:

    @app.get("/")
    @login_required
    def index():
        return render_template("index.html", terceros_cache=directorio.contar())


    @app.post("/cargar")
    @login_required
    def cargar():
        archivo = request.files.get("archivo")
        if not archivo or archivo.filename == "":
            flash("Selecciona un archivo XLSX.", "error")
            return redirect(url_for("index"))

        nombre = secure_filename(archivo.filename)
        destino = _user_dir("") / nombre
        archivo.save(destino)

        try:
            df, informantes = parser.cargar_archivo(str(destino))
        except Exception as exc:
            flash(f"Error leyendo el archivo: {exc}", "error")
            return redirect(url_for("index"))

        _guardar_df(df)

        no_mapeados = parser.detectar_tipos_no_mapeados(df, registry.tipos_conocidos())
        session["no_mapeados"] = no_mapeados
        session["informantes"] = [{"nit": i.nit, "nombre": i.nombre} for i in informantes]
        session["archivo_nombre"] = nombre

        if no_mapeados:
            return redirect(url_for("conceptos"))
        return redirect(url_for("preview"))


    @app.get("/conceptos")
    @login_required
    def conceptos():
        no_mapeados = session.get("no_mapeados", [])
        if not no_mapeados:
            return redirect(url_for("preview"))
        return render_template("conceptos.html", tipos=no_mapeados, categorias=registry.CATEGORIAS_VALIDAS)


    @app.post("/conceptos")
    @login_required
    def conceptos_post():
        no_mapeados = session.get("no_mapeados", [])
        for tipo in no_mapeados:
            cat = request.form.get(f"cat_{tipo}", "ignorar")
            signo = int(request.form.get(f"signo_{tipo}", "1"))
            registry.registrar_tipo(tipo, cat, signo)
        session["no_mapeados"] = []
        return redirect(url_for("preview"))


    @app.get("/preview")
    @login_required
    def preview():
        df = _cargar_df()
        if df is None:
            return redirect(url_for("index"))
        resumen = {
            "filas_totales": len(df),
            "por_tipo": df["tipo_documento"].value_counts().to_dict() if "tipo_documento" in df.columns else {},
            "por_grupo": df["grupo"].value_counts().to_dict() if "grupo" in df.columns else {},
            "rango_fechas": (
                (df["fecha_emision"].min().strftime("%Y-%m-%d") if df["fecha_emision"].notna().any() else "—"),
                (df["fecha_emision"].max().strftime("%Y-%m-%d") if df["fecha_emision"].notna().any() else "—"),
            ) if "fecha_emision" in df.columns else ("—", "—"),
        }
        return render_template(
            "preview.html",
            informante=_informante_actual(),
            archivo=session.get("archivo_nombre", ""),
            resumen=resumen,
            concepto_1001=registry.concepto_default("1001") or 5016,
            concepto_1007=registry.concepto_default("1007") or 4001,
            concepto_5248=4010,
        )


    @app.post("/generar")
    @login_required
    def generar():
        df = _cargar_df()
        if df is None:
            return redirect(url_for("index"))

        informante = _informante_actual()
        if not informante:
            flash("No se detectó el NIT del informante en el XLSX. Revisa el archivo cargado.", "error")
            return redirect(url_for("preview"))

        modo_colaboracion = request.form.get("modo_colaboracion", "")
        if modo_colaboracion not in ("si", "no"):
            flash("Indica obligatoriamente si es consorcio / contrato de colaboración.", "error")
            return redirect(url_for("preview"))

        nit_inf = informante["nit"]
        nombre_inf = informante["nombre"]
        ano = _entero_form(request.form.get("ano_gravable"), 2025)
        mes_ini = _entero_form(request.form.get("mes_inicio"), 1)
        mes_fin = _entero_form(request.form.get("mes_fin"), 12)
        cpt_1001 = _entero_form(request.form.get("cpt_1001"), registry.concepto_default("1001") or 5016)
        es_part = modo_colaboracion == "si"
        if es_part:
            cpt_1007 = _entero_form(request.form.get("cpt_5248"), 4010)
            tcon = _entero_form(request.form.get("tipo_contrato"), 2)
            nit_part = nit_inf
            tdoc_part = helpers.inferir_tipo_documento(nit_inf, nombre_inf)
        else:
            cpt_1007 = _entero_form(request.form.get("cpt_1007"), registry.concepto_default("1007") or 4001)
            tcon = 2
            nit_part = ""
            tdoc_part = 31
        idfi = request.form.get("id_fideicomiso", "").strip()

        ctx = ContextoInformante(
            nit=nit_inf, razon_social=nombre_inf, ano_gravable=ano,
            mes_inicio=mes_ini, mes_fin=mes_fin,
            concepto_default_1001=cpt_1001, concepto_default_1007=cpt_1007,
            es_participante_colaboracion=es_part, tipo_contrato=tcon,
            nit_participante=nit_part, tdoc_participante=tdoc_part,
            id_fideicomiso=idfi,
        )

        df_filtrado = parser.filtrar_por_periodo(df, ano, mes_ini, mes_fin)
        df_generacion = preparar_dataframe_generacion(df_filtrado, ctx)

        dataframes = {}
        formatos_objetivo = (
            ("5247", "5248", "5249", "5250") if es_part
            else ("1001", "1005", "1006", "1007")
        )
        for codigo in formatos_objetivo:
            GenClass = GENERADORES[codigo]
            gen = GenClass(ctx)
            dataframes[codigo] = gen.generar(df_generacion)

        nit_archivo = secure_filename(nit_inf) or "INFORMANTE"
        nombre_archivo = f"Exogena_AG{ano}_{nit_archivo}.xlsx"
        archivo_origen_nombre = session.get("archivo_nombre", "")
        archivo_origen = (_user_dir("") / archivo_origen_nombre) if archivo_origen_nombre else None
        archivo = _crear_xlsx_consolidado(dataframes, ctx, df_filtrado, archivo_origen, df_generacion)
        return send_file(
            archivo,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


    @app.get("/directorio")
    @login_required
    def ver_directorio():
        data = directorio.listar()
        df = _cargar_df()
        faltantes = directorio.nits_faltantes_del_df(df) if df is not None else []
        return render_template("directorio.html", terceros=data, total=len(data),
                               faltantes=faltantes, tiene_df=df is not None)


    @app.post("/directorio/importar")
    @login_required
    def importar_directorio():
        archivos = request.files.getlist("xmls")
        if not archivos:
            flash("Selecciona uno o más XML / ZIP.", "error")
            return redirect(url_for("ver_directorio"))
        tmp = Path(tempfile.mkdtemp(prefix="ubl_"))
        try:
            for f in archivos:
                raw = f.filename or ""
                partes = [secure_filename(p) for p in raw.replace("\\", "/").split("/") if p]
                if not partes:
                    continue
                destino = tmp.joinpath(*partes)
                destino.parent.mkdir(parents=True, exist_ok=True)
                f.save(destino)
            leidos, actualizados = directorio.importar_carpeta(tmp)
            flash(f"Importados {actualizados} registros de tercero desde {leidos} XML.", "ok")
        finally:
            shutil.rmtree(tmp, ignore_errors=True)
        return redirect(url_for("ver_directorio"))


    @app.post("/directorio/importar_ruta")
    @login_required
    def importar_ruta_local():
        # Solo permitir esta función si estás en localhost (no en Render)
        if os.environ.get("PERMITIR_RUTA_LOCAL", "false").lower() != "true":
            flash("Importación por ruta local deshabilitada en este entorno. Usa subida de archivos.", "error")
            return redirect(url_for("ver_directorio"))
        ruta = (request.form.get("ruta") or "").strip().strip('"')
        if not ruta:
            flash("Indica la ruta de la carpeta.", "error")
            return redirect(url_for("ver_directorio"))
        carpeta = Path(ruta)
        if not carpeta.exists() or not carpeta.is_dir():
            flash(f"La ruta no existe: {ruta}", "error")
            return redirect(url_for("ver_directorio"))
        leidos, actualizados = directorio.importar_carpeta(carpeta)
        flash(f"Importados {actualizados} registros desde {leidos} XML en {ruta}", "ok")
        return redirect(url_for("ver_directorio"))


    @app.post("/directorio/limpiar")
    @login_required
    def limpiar_directorio():
        directorio.limpiar()
        flash("Directorio vacío.", "ok")
        return redirect(url_for("ver_directorio"))


    @app.post("/directorio/enriquecer")
    @login_required
    def enriquecer_directorio():
        df = _cargar_df()
        if df is None:
            flash("Sube primero el XLSX MUISCA.", "error")
            return redirect(url_for("ver_directorio"))
        faltantes = directorio.nits_faltantes_del_df(df)
        if not faltantes:
            flash("No hay NITs faltantes.", "ok")
            return redirect(url_for("ver_directorio"))
        forzar = request.form.get("forzar") == "on"
        enc, noenc, errores = directorio.enriquecer_lote(faltantes, forzar=forzar, delay=0.3)
        msg = f"RUES: {enc} encontrados, {noenc} no encontrados"
        if errores:
            msg += f", {len(errores)} con error"
        flash(msg, "ok" if enc else "error")
        return redirect(url_for("ver_directorio"))


    @app.post("/directorio/recalcular_nombres")
    @login_required
    def recalcular_nombres_directorio():
        df = _cargar_df()
        if df is None:
            flash("Sube primero el XLSX MUISCA.", "error")
            return redirect(url_for("ver_directorio"))
        actualizados = directorio.recalcular_nombres_desde_df(df)
        flash(f"Nombres recalculados para {actualizados} personas naturales.", "ok")
        return redirect(url_for("ver_directorio"))


    @app.post("/directorio/editar")
    @login_required
    def editar_tercero():
        nid = request.form.get("nid", "").strip()
        if not nid:
            flash("NIT requerido.", "error")
            return redirect(url_for("ver_directorio"))
        datos = {
            "raz": request.form.get("raz", "").strip(),
            "apl1": request.form.get("apl1", "").strip(),
            "apl2": request.form.get("apl2", "").strip(),
            "nom1": request.form.get("nom1", "").strip(),
            "nom2": request.form.get("nom2", "").strip(),
            "dir": request.form.get("dir", "").strip(),
            "dpto": request.form.get("dpto", "").strip(),
            "mun": request.form.get("mun", "").strip(),
            "pais": int(request.form.get("pais") or 169),
        }
        dv_str = request.form.get("dv", "").strip()
        if dv_str.isdigit():
            datos["dv"] = int(dv_str)
        tdoc = _entero_form(request.form.get("tdoc"), 0)
        if tdoc:
            datos["tdoc"] = tdoc
        datos = {k: v for k, v in datos.items() if v not in ("", None)}
        directorio.actualizar_manual(nid, datos)
        flash(f"Tercero {nid} actualizado.", "ok")
        return redirect(url_for("ver_directorio"))


    @app.post("/directorio/eliminar")
    @login_required
    def eliminar_tercero():
        nid = request.form.get("nid", "").strip()
        if nid:
            directorio.eliminar(nid)
            flash(f"Tercero {nid} eliminado.", "ok")
        return redirect(url_for("ver_directorio"))


# ---------------- exportación XLSX ----------------


def _crear_xlsx_consolidado(
    dataframes: dict[str, pd.DataFrame],
    ctx: ContextoInformante,
    df_origen: pd.DataFrame,
    archivo_origen: Path | None = None,
    df_generacion: pd.DataFrame | None = None,
) -> io.IOBase:
    if _requiere_xlsx_streaming(df_origen, archivo_origen):
        return _crear_xlsx_consolidado_streaming(
            dataframes, ctx, df_origen, archivo_origen, df_generacion
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    _llenar_resumen(ws, df_generacion if df_generacion is not None else df_origen, ctx)
    ws = wb.create_sheet(title="Informe")
    if not _copiar_primera_hoja_cargada(archivo_origen, ws):
        _llenar_informe_normalizado(ws, df_origen)
    ws = wb.create_sheet(title="Índice")
    ws.cell(1, 1, value=f"Exógena DIAN AG{ctx.ano_gravable}").font = Font(bold=True, size=14)
    ws.cell(2, 1, value=f"Informante: {ctx.nit} - {ctx.razon_social}")
    ws.cell(3, 1, value=f"Periodo: meses {ctx.mes_inicio:02d} a {ctx.mes_fin:02d}")
    ws.cell(5, 1, value="Formato").font = Font(bold=True)
    ws.cell(5, 2, value="Descripción").font = Font(bold=True)
    ws.cell(5, 3, value="Registros").font = Font(bold=True)
    cfg = _config()["formatos"]
    fila = 6
    for codigo, df in dataframes.items():
        ws.cell(fila, 1, value=codigo)
        ws.cell(fila, 2, value=cfg[codigo]["nombre"])
        ws.cell(fila, 3, value=len(df))
        fila += 1
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12
    for codigo, df in dataframes.items():
        hoja = wb.create_sheet(title=f"F{codigo}")
        _llenar_hoja(hoja, df, codigo, ctx, with_meta=False)
    archivo = _archivo_temporal_xlsx()
    wb.save(archivo)
    archivo.seek(0)
    return archivo


def _archivo_temporal_xlsx():
    return tempfile.SpooledTemporaryFile(max_size=XLSX_SPOOL_MAX_SIZE, mode="w+b")


def _requiere_xlsx_streaming(df_origen: pd.DataFrame, archivo_origen: Path | None) -> bool:
    celdas_df = len(df_origen) * max(1, len(df_origen.columns))
    if celdas_df >= XLSX_STREAMING_CELL_THRESHOLD:
        return True
    if not archivo_origen or not archivo_origen.exists():
        return False
    try:
        wb = load_workbook(archivo_origen, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        celdas_origen = (ws.max_row or 0) * max(1, ws.max_column or 1)
        return celdas_origen >= XLSX_STREAMING_CELL_THRESHOLD
    except Exception:
        return False
    finally:
        if "wb" in locals():
            wb.close()


def _crear_xlsx_consolidado_streaming(
    dataframes: dict[str, pd.DataFrame],
    ctx: ContextoInformante,
    df_origen: pd.DataFrame,
    archivo_origen: Path | None = None,
    df_generacion: pd.DataFrame | None = None,
) -> io.IOBase:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Resumen")
    _llenar_resumen_streaming(ws, df_generacion if df_generacion is not None else df_origen, ctx)

    ws = wb.create_sheet(title="Informe")
    if not _copiar_primera_hoja_cargada_streaming(archivo_origen, ws):
        _llenar_informe_normalizado_streaming(ws, df_origen)

    ws = wb.create_sheet(title="Índice")
    _llenar_indice_streaming(ws, dataframes, ctx)

    for codigo, df in dataframes.items():
        hoja = wb.create_sheet(title=f"F{codigo}")
        _llenar_hoja_streaming(hoja, df, codigo, ctx, with_meta=False)

    archivo = _archivo_temporal_xlsx()
    wb.save(archivo)
    archivo.seek(0)
    return archivo


def _resumen_empresa(df: pd.DataFrame, ctx: ContextoInformante) -> dict[str, float]:
    formato_ingresos = "5248" if ctx.es_participante_colaboracion else "1007"
    formato_gastos = "5247" if ctx.es_participante_colaboracion else "1001"
    reglas_ingresos = registry.regla_para_formato(formato_ingresos)
    reglas_gastos = registry.regla_para_formato(formato_gastos)
    out = {
        "ingresos_brutos": 0.0,
        "devoluciones_ingresos": 0.0,
        "iva_generado": 0.0,
        "total_ingresos": 0.0,
        "gastos_brutos": 0.0,
        "devoluciones_gastos": 0.0,
        "iva_descontable": 0.0,
        "total_gastos": 0.0,
        "nomina": 0.0,
    }

    if df.empty:
        out["ingresos_netos"] = 0.0
        out["gastos_netos"] = 0.0
        out["utilidad_estimada"] = 0.0
        return out

    work = df if "__categoria__" in df.columns and "__signo__" in df.columns else preparar_dataframe_generacion(df, ctx)
    categorias = work["__categoria__"].fillna("").astype(str)
    grupos = work.get("__grupo__", work["grupo"]).fillna("").astype(str).str.strip()
    signos = pd.to_numeric(work["__signo__"], errors="coerce").fillna(1.0)
    zeros = pd.Series(0.0, index=work.index)
    base = pd.to_numeric(work["base"] if "base" in work.columns else zeros, errors="coerce").fillna(0.0) * signos
    iva = pd.to_numeric(work["iva"] if "iva" in work.columns else zeros, errors="coerce").fillna(0.0) * signos
    total = pd.to_numeric(work["total"] if "total" in work.columns else zeros, errors="coerce").fillna(0.0) * signos

    nomina_mask = categorias.eq("nomina")
    out["nomina"] = float(base.where(nomina_mask & (base > 0), 0.0).sum())

    reglas_ingresos_idx = pd.MultiIndex.from_tuples(reglas_ingresos)
    reglas_gastos_idx = pd.MultiIndex.from_tuples(reglas_gastos)
    pares = pd.MultiIndex.from_arrays([categorias, grupos])
    ingresos_mask = pd.Series(pares.isin(reglas_ingresos_idx), index=work.index)
    gastos_mask = pd.Series(pares.isin(reglas_gastos_idx), index=work.index)

    out["ingresos_brutos"] = float(base.where(ingresos_mask & (base >= 0), 0.0).sum())
    out["devoluciones_ingresos"] = float((-base.where(ingresos_mask & (base < 0), 0.0)).sum())
    out["iva_generado"] = float(iva.where(ingresos_mask, 0.0).sum())
    out["total_ingresos"] = float(total.where(ingresos_mask, 0.0).sum())
    out["gastos_brutos"] = float(base.where(gastos_mask & (base >= 0), 0.0).sum())
    out["devoluciones_gastos"] = float((-base.where(gastos_mask & (base < 0), 0.0)).sum())
    out["iva_descontable"] = float(iva.where(gastos_mask, 0.0).sum())
    out["total_gastos"] = float(total.where(gastos_mask, 0.0).sum())

    out["ingresos_netos"] = out["ingresos_brutos"] - out["devoluciones_ingresos"]
    out["gastos_netos"] = out["gastos_brutos"] - out["devoluciones_gastos"]
    out["utilidad_estimada"] = out["ingresos_netos"] - out["gastos_netos"] - out["nomina"]
    return out


def _valor_resumen(ws, fila: int, etiqueta: str, valor: float | None, *, fuerte: bool = False) -> None:
    ws.cell(fila, 1, value=etiqueta)
    ws.cell(fila, 2, value="$" if valor is not None else "")
    ws.cell(fila, 3, value="-" if valor in (None, 0) else round(valor))
    for col in range(1, 4):
        celda = ws.cell(fila, col)
        celda.border = Border(bottom=Side(style="thin", color="D1D5DB"))
        celda.alignment = Alignment(vertical="center")
        if fuerte:
            celda.font = Font(bold=True)
            celda.fill = PatternFill("solid", fgColor="F3F4F6")
    ws.cell(fila, 2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(fila, 3).alignment = Alignment(horizontal="right", vertical="center")
    if valor not in (None, 0):
        ws.cell(fila, 3).number_format = '#,##0;[Red]-#,##0;"-"'


def _seccion_resumen(ws, fila: int, titulo: str) -> None:
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
    celda = ws.cell(fila, 1, value=titulo)
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill("solid", fgColor="1E3A8A")
    celda.alignment = Alignment(horizontal="left", vertical="center")


def _llenar_resumen(ws, df: pd.DataFrame, ctx: ContextoInformante) -> None:
    resumen = _resumen_empresa(df, ctx)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:C1")
    ws.cell(1, 1, value="Resumen financiero").font = Font(bold=True, size=18, color="1E3A8A")
    ws.cell(2, 1, value=f"{ctx.razon_social} · NIT {ctx.nit}")
    ws.cell(3, 1, value=f"Año gravable {ctx.ano_gravable} · Meses {ctx.mes_inicio:02d} a {ctx.mes_fin:02d}")
    ws.cell(4, 1, value="Valores estimados desde el XLSX MUISCA cargado. Utilidad calculada sobre bases sin IVA.")
    ws.cell(4, 1).font = Font(italic=True, color="6B7280")

    fila = 6
    _seccion_resumen(ws, fila, "Ingresos")
    fila += 1
    _valor_resumen(ws, fila, "Ingresos brutos", resumen["ingresos_brutos"])
    fila += 1
    _valor_resumen(ws, fila, "Devoluciones, rebajas y descuentos", resumen["devoluciones_ingresos"])
    fila += 1
    _valor_resumen(ws, fila, "Ingresos netos", resumen["ingresos_netos"], fuerte=True)
    fila += 1
    _valor_resumen(ws, fila, "IVA generado neto", resumen["iva_generado"])
    fila += 1
    _valor_resumen(ws, fila, "Total ingresos con IVA", resumen["total_ingresos"], fuerte=True)

    fila += 2
    _seccion_resumen(ws, fila, "Gastos y costos")
    fila += 1
    _valor_resumen(ws, fila, "Gastos / costos brutos", resumen["gastos_brutos"])
    fila += 1
    _valor_resumen(ws, fila, "Devoluciones en gastos", resumen["devoluciones_gastos"])
    fila += 1
    _valor_resumen(ws, fila, "Gastos / costos netos", resumen["gastos_netos"], fuerte=True)
    fila += 1
    _valor_resumen(ws, fila, "IVA descontable / mayor valor neto", resumen["iva_descontable"])
    fila += 1
    _valor_resumen(ws, fila, "Total gastos con IVA", resumen["total_gastos"], fuerte=True)

    fila += 2
    _seccion_resumen(ws, fila, "Nómina")
    fila += 1
    _valor_resumen(ws, fila, "Nómina identificada", resumen["nomina"], fuerte=True)

    fila += 2
    _seccion_resumen(ws, fila, "Resultado estimado")
    fila += 1
    _valor_resumen(ws, fila, "Utilidad estimada sin IVA", resumen["utilidad_estimada"], fuerte=True)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 18
    for row in range(1, fila + 1):
        ws.row_dimensions[row].height = 24


def _celda_streaming(
    ws,
    valor,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
    number_format: str | None = None,
):
    celda = WriteOnlyCell(ws, value=_valor_excel(valor))
    if font is not None:
        celda.font = font
    if fill is not None:
        celda.fill = fill
    if alignment is not None:
        celda.alignment = alignment
    if border is not None:
        celda.border = border
    if number_format is not None:
        celda.number_format = number_format
    return celda


def _append_streaming(
    ws,
    valores,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
):
    ws.append([
        _celda_streaming(ws, valor, font=font, fill=fill, alignment=alignment, border=border)
        for valor in valores
    ])


def _nombres_encabezado_unicos(valores) -> list[str]:
    vistos: dict[str, int] = {}
    nombres: list[str] = []
    for idx, valor in enumerate(valores, 1):
        base = str(valor).strip() if valor is not None else ""
        if not base:
            base = f"Columna {idx}"
        veces = vistos.get(base, 0) + 1
        vistos[base] = veces
        nombres.append(base if veces == 1 else f"{base} {veces}")
    return nombres


def _ajustar_columnas_streaming(ws, encabezados: list[str]) -> None:
    for idx, encabezado in enumerate(encabezados, 1):
        letra = get_column_letter(idx)
        ws.column_dimensions[letra].width = max(10, min(42, len(str(encabezado)) + 2))


def _agregar_tabla_streaming(ws, nombre: str, encabezados: list[str], max_row: int, max_col: int) -> None:
    if max_row < 1 or max_col < 1:
        return
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.auto_filter.ref = ref
    if max_row <= 1:
        return
    try:
        tabla = Table(displayName=nombre, ref=ref)
        tabla._initialise_columns()
        for columna, encabezado in zip(tabla.tableColumns, encabezados[:max_col]):
            columna.name = str(encabezado)
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="In write-only mode you must add table columns manually",
                category=UserWarning,
            )
            ws.add_table(tabla)
    except Exception:
        ws.auto_filter.ref = ref


def _llenar_resumen_streaming(ws, df: pd.DataFrame, ctx: ContextoInformante) -> None:
    resumen = _resumen_empresa(df, ctx)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 18

    titulo_font = Font(bold=True, size=16, color="1E3A8A")
    subtitulo_font = Font(color="374151")
    header_font = Font(bold=True, color="111827")
    header_fill = PatternFill("solid", fgColor="F3F4F6")

    _append_streaming(ws, ["Resumen financiero"], font=titulo_font)
    _append_streaming(ws, [f"{ctx.razon_social} · NIT {ctx.nit}"], font=subtitulo_font)
    _append_streaming(ws, [f"Año gravable {ctx.ano_gravable} · Meses {ctx.mes_inicio:02d} a {ctx.mes_fin:02d}"])
    ws.append([])
    _append_streaming(ws, ["Sección", "Concepto", "Moneda", "Valor"], font=header_font, fill=header_fill)

    filas = [
        ("Ingresos", "Ingresos brutos", resumen["ingresos_brutos"]),
        ("Ingresos", "Devoluciones, rebajas y descuentos", resumen["devoluciones_ingresos"]),
        ("Ingresos", "Ingresos netos", resumen["ingresos_netos"]),
        ("Ingresos", "IVA generado neto", resumen["iva_generado"]),
        ("Ingresos", "Total ingresos con IVA", resumen["total_ingresos"]),
        ("Gastos y costos", "Gastos / costos brutos", resumen["gastos_brutos"]),
        ("Gastos y costos", "Devoluciones en gastos", resumen["devoluciones_gastos"]),
        ("Gastos y costos", "Gastos / costos netos", resumen["gastos_netos"]),
        ("Gastos y costos", "IVA descontable / mayor valor neto", resumen["iva_descontable"]),
        ("Gastos y costos", "Total gastos con IVA", resumen["total_gastos"]),
        ("Nómina", "Nómina identificada", resumen["nomina"]),
        ("Resultado estimado", "Utilidad estimada sin IVA", resumen["utilidad_estimada"]),
    ]
    for seccion, concepto, valor in filas:
        numero = "-" if valor in (None, 0) else round(valor)
        row = [
            _celda_streaming(ws, seccion),
            _celda_streaming(ws, concepto),
            _celda_streaming(ws, "$" if valor not in (None, 0) else "", alignment=Alignment(horizontal="center")),
            _celda_streaming(ws, numero, alignment=Alignment(horizontal="right"), number_format='#,##0;[Red]-#,##0;"-"'),
        ]
        ws.append(row)


def _llenar_indice_streaming(
    ws,
    dataframes: dict[str, pd.DataFrame],
    ctx: ContextoInformante,
) -> None:
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 12
    titulo_font = Font(bold=True, size=14)
    bold = Font(bold=True)
    _append_streaming(ws, [f"Exógena DIAN AG{ctx.ano_gravable}"], font=titulo_font)
    ws.append([f"Informante: {ctx.nit} - {ctx.razon_social}"])
    ws.append([f"Periodo: meses {ctx.mes_inicio:02d} a {ctx.mes_fin:02d}"])
    ws.append([])
    _append_streaming(ws, ["Formato", "Descripción", "Registros"], font=bold)
    cfg = _config()["formatos"]
    for codigo, df in dataframes.items():
        ws.append([codigo, cfg[codigo]["nombre"], len(df)])


def _copiar_primera_hoja_cargada_streaming(archivo_origen: Path | None, destino) -> bool:
    if not archivo_origen or not archivo_origen.exists():
        return False
    try:
        wb_origen = load_workbook(archivo_origen, read_only=True, data_only=False)
    except Exception:
        return False

    try:
        origen = wb_origen.worksheets[0]
        try:
            min_col, min_row, max_col, max_row = range_boundaries(origen.calculate_dimension())
        except Exception:
            min_row, max_row = 1, origen.max_row or 0
            min_col, max_col = 1, origen.max_column or 0
        if max_row < 1 or max_col < 1:
            return False

        destino.sheet_view.showGridLines = False
        destino.freeze_panes = "A2" if max_row > 1 else None
        header_font = Font(bold=True, size=10, color="111827")
        header_fill = PatternFill("solid", fgColor="F3F4F6")
        header_border = Border(bottom=Side(style="thin", color="D1D5DB"))
        encabezados: list[str] = []
        filas_escritas = 0

        for row_idx, fila in enumerate(
            origen.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col),
            1,
        ):
            valores = [celda.value for celda in fila]
            if row_idx == 1:
                encabezados = _nombres_encabezado_unicos(valores)
                _append_streaming(
                    destino,
                    encabezados,
                    font=header_font,
                    fill=header_fill,
                    border=header_border,
                    alignment=Alignment(vertical="center", wrap_text=False),
                )
            else:
                destino.append([_valor_excel(valor) for valor in valores])
            filas_escritas += 1

        _ajustar_columnas_streaming(destino, encabezados)
        _agregar_tabla_streaming(
            destino,
            "TablaInformeCargado",
            encabezados,
            filas_escritas,
            max_col - min_col + 1,
        )
        return filas_escritas > 0
    finally:
        wb_origen.close()


def _llenar_informe_normalizado_streaming(ws, df: pd.DataFrame) -> None:
    columnas = _nombres_encabezado_unicos(list(df.columns))
    if not columnas:
        return
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2" if len(df) else None
    header_font = Font(bold=True, size=10, color="111827")
    header_fill = PatternFill("solid", fgColor="F3F4F6")
    _append_streaming(ws, columnas, font=header_font, fill=header_fill)
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append([_valor_excel(valor) for valor in row])
    _ajustar_columnas_streaming(ws, columnas)
    _agregar_tabla_streaming(ws, "TablaInformeCargado", columnas, len(df) + 1, len(columnas))


def _llenar_hoja_streaming(
    ws,
    df: pd.DataFrame,
    codigo: str,
    ctx: ContextoInformante,
    with_meta: bool,
):
    cfg = _config()["formatos"][codigo]
    columnas = cfg["columnas"]
    if with_meta:
        _append_streaming(ws, [f"Formato {codigo} v{cfg['version']} - {cfg['nombre']}"], font=Font(bold=True, size=12))
        ws.append([f"Informante: {ctx.nit} - {ctx.razon_social}"])
        ws.append([f"Periodo: {ctx.ano_gravable} ({ctx.mes_inicio:02d} a {ctx.mes_fin:02d})"])
        ws.append([f"Registros: {len(df)}"])
        ws.append([])

    fill = PatternFill("solid", fgColor="FFE4B5")
    bold = Font(bold=True)
    italic = Font(italic=True, color="808080")
    _append_streaming(ws, [c["nombre"] for c in columnas], font=bold, fill=fill)
    _append_streaming(ws, [c["campo"] for c in columnas], font=italic)
    campos = [c["campo"] for c in columnas]
    for valores in df[campos].itertuples(index=False, name=None):
        ws.append([_valor_excel(valor) for valor in valores])
    for idx, c in enumerate(columnas, 1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(40, len(c["nombre"]) + 2))


def _copiar_primera_hoja_cargada(archivo_origen: Path | None, destino) -> bool:
    if not archivo_origen or not archivo_origen.exists():
        return False
    try:
        wb_origen = load_workbook(archivo_origen, data_only=False)
    except Exception:
        return False

    try:
        origen = wb_origen.worksheets[0]
        limites = _rango_con_contenido(origen)
        if not limites:
            return False

        min_row, max_row, min_col, max_col = limites
        for src_row in range(min_row, max_row + 1):
            dst_row = src_row - min_row + 1
            for src_col in range(min_col, max_col + 1):
                dst_col = src_col - min_col + 1
                origen_celda = origen.cell(src_row, src_col)
                destino_celda = destino.cell(dst_row, dst_col, value=origen_celda.value)
                destino_celda.number_format = origen_celda.number_format

        _formatear_tabla_informe(destino, max_row - min_row + 1, max_col - min_col + 1)
        return True
    finally:
        wb_origen.close()


def _llenar_informe_normalizado(ws, df: pd.DataFrame) -> None:
    columnas = list(df.columns)
    if not columnas:
        return

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append([_valor_excel(valor) for valor in row])
    _formatear_tabla_informe(ws, len(df) + 1, len(columnas))


def _rango_con_contenido(ws) -> tuple[int, int, int, int] | None:
    min_row = min_col = None
    max_row = max_col = 0
    for fila in ws.iter_rows():
        for celda in fila:
            valor = celda.value
            if valor is None or (isinstance(valor, str) and not valor.strip()):
                continue
            min_row = celda.row if min_row is None else min(min_row, celda.row)
            min_col = celda.column if min_col is None else min(min_col, celda.column)
            max_row = max(max_row, celda.row)
            max_col = max(max_col, celda.column)
    if min_row is None or min_col is None:
        return None
    return min_row, max_row, min_col, max_col


def _valor_excel(valor):
    if valor is None:
        return None
    try:
        return None if pd.isna(valor) else valor
    except (TypeError, ValueError):
        return valor


def _formatear_tabla_informe(ws, max_row: int, max_col: int) -> None:
    if max_row < 1 or max_col < 1:
        return

    _asegurar_encabezados_tabla(ws, max_col)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2" if max_row > 1 else None

    encabezado_fill = PatternFill("solid", fgColor="F3F4F6")
    encabezado_font = Font(bold=True, size=10, color="111827")
    ws.row_dimensions[1].height = 22
    for col_idx in range(1, max_col + 1):
        celda = ws.cell(1, col_idx)
        celda.border = Border(bottom=Side(style="thin", color="D1D5DB"))
        celda.alignment = Alignment(vertical="center", wrap_text=False)
        celda.font = encabezado_font
        celda.fill = encabezado_fill

    for col_idx in range(1, max_col + 1):
        letra = get_column_letter(col_idx)
        muestras = [ws.cell(row_idx, col_idx).value for row_idx in range(1, min(max_row, 80) + 1)]
        ancho = max(len(str(valor)) if valor is not None else 0 for valor in muestras)
        ws.column_dimensions[letra].width = max(10, min(42, ancho + 2))

    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    if max_row > 1:
        tabla = Table(displayName="TablaInformeCargado", ref=ref)
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight1",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tabla)
    else:
        ws.auto_filter.ref = ref


def _asegurar_encabezados_tabla(ws, max_col: int) -> None:
    vistos: dict[str, int] = {}
    for col_idx in range(1, max_col + 1):
        celda = ws.cell(1, col_idx)
        nombre = str(celda.value).strip() if celda.value is not None else ""
        if not nombre:
            nombre = f"Columna {col_idx}"
        veces = vistos.get(nombre, 0)
        vistos[nombre] = veces + 1
        if veces:
            nombre = f"{nombre} {veces + 1}"
        celda.value = nombre


def _llenar_hoja(ws, df: pd.DataFrame, codigo: str, ctx: ContextoInformante, with_meta: bool):
    cfg = _config()["formatos"][codigo]
    columnas = cfg["columnas"]
    fila_header = 1
    if with_meta:
        ws.cell(1, 1, value=f"Formato {codigo} v{cfg['version']} - {cfg['nombre']}").font = Font(bold=True, size=12)
        ws.cell(2, 1, value=f"Informante: {ctx.nit} - {ctx.razon_social}")
        ws.cell(3, 1, value=f"Periodo: {ctx.ano_gravable} ({ctx.mes_inicio:02d} a {ctx.mes_fin:02d})")
        ws.cell(4, 1, value=f"Registros: {len(df)}")
        fila_header = 6
    fill = PatternFill("solid", fgColor="FFE4B5")
    bold = Font(bold=True)
    italic = Font(italic=True, color="808080")
    for j, c in enumerate(columnas, 1):
        cel = ws.cell(fila_header, j, value=c["nombre"])
        cel.font = bold
        cel.fill = fill
        ws.cell(fila_header + 1, j, value=c["campo"]).font = italic
    for row in dataframe_to_rows(df[[c["campo"] for c in columnas]], index=False, header=False):
        ws.append([_valor_excel(valor) for valor in row])
    for j, c in enumerate(columnas, 1):
        ws.column_dimensions[ws.cell(1, j).column_letter].width = max(12, min(40, len(c["nombre"]) + 2))


app = crear_app()


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
